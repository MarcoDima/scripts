# python -m pip install -U pip setuptools for Windows
# pip install -U pip setuptools For Linux/OS X
#  

from openpyxl import Workbook

from openpyxl import load_workbook
import os
import pandas as pd
import sys

# Retrieve current working directory (`cwd`)
cwd = os.getcwd()

#print ("Current working directory: " + cwd)
# Change directory 
os.chdir("/home/dima/python/example")

# List all files and directories in current directory
#print("Listing all files in the dir:")
#print(os.listdir('.'))
pathToFile="/to/path/file"

def findAllQuotas(ws, c):
    quotas=[]
    start_l = c
    start_num = "4"
    start = start_l + start_num
    cell = ws[start]
    while not cell.value is None:
        quotas.append(cell.value)
        start_num = str(int(start_num)+1)
        start = start_l + start_num
        cell = ws[start]

    #for q in ws.iter_rows(min_row=1, max_col=30, max_row=2):
    
    return quotas

def printSemiColonValues(list):
    f = open('/home/dima/python/example/log.txt', 'ab')
    for value in list:
        f.write(value + ";")
        #sys.stdout.write(value + ";")
    sys.stdout.flush()

### Definitions of array of quotas

wb = load_workbook('example.xlsx')
ws = wb.active

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell.value)

### print quotas ##########
for q in findAllQuotas(ws, "M"):
    sys.stdout.write(q+ " ")

for q1 in findAllQuotas(ws, "N"):
    sys.stdout.write(str(q1)+ " ")

###########################


st = list("Hello World")
st[2] = "L"
print("".join(st))
#d = ws.cell(row=4, column=2, value=10)
#print(ws["M5"].value)
#print wb.get_sheet_names()
#wb = open_workbook('example.xlsx')
#for s in wb.sheets():
 #   values = []
  #  for row in range(s.nrows):
     #   col_value = []
   #     for col in range(s.ncols):
    #        value  = (s.cell(row,col).value)
      #      try : value = str(int(value))
       #     except : pass
        #    col_value.append(value)
#        values.append(col_value)
#print values

#for el in values:
#    for el1 in el:
#        for q in quotas:
#           if (el1 == q):
#                print ("Found quotas")
#            else:
#               print("No quotas found")


#printSemiColonValues(values[0])
#print("")
#printSemiColonValues(values[1])


#print(values[0][0]) 
#print(", ")
#print(values[0][1])
#print(s.nrows)